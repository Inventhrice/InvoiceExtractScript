import openpyxl as opyxl

dir = "" #DIR

DLEmailList = opyxl.load_workbook((dir+"DLEmailList.xlsx"))
sheet = DLEmailList.active

writeFile = open("MissedInvoicesList.txt", 'w')
with open((dir+"CSBInvoice.txt")) as CSBdata:
    for allData in enumerate(CSBdata):
        i = 2
        found = False
        line = allData[1]
        data = line.split(';')
        while i < sheet.max_row+1 and (found == False):
            
            if (str(data[2]) == str(sheet.cell(row=i, column=1).value)):
                found = True
            i+=1
        if not found:
            writeFile.write(line)
writeFile.close()



#sheet = DLEmailList.active

# vvv this used to be the VBA code. It didnt work out because we couldnt figure out how to call the macro without multiple problems.
""" Sub Workbook_open()

Dim Wb1 As Workbook
Dim found As Variant
Dim FName As Variant
Dim i As Integer

' Open the other workbook
' Input the FULL path to the file, including its extension
ThisWorkbook.Connections(1).OLEDBConnection.Refresh

Set Wb1 = Workbooks.Open("\\\cmpfs01\Departmental\Information Systems\Private\InvoiceExtract\DLEmailList.xlsx")

FName = "\\cmpfs01\Departmental\Information Systems\Private\InvoiceExtract\MissedInvoicesList.txt"

Open FName For Output As #1
Print #1, "Missed orders:"

For i = 2 To Worksheets(1).UsedRange.Rows.Count
     found = Application.Match(Worksheets(1).Cells(i, 3).Value, Wb1.Sheets(1).Range("A:A"), 0)
     
     If IsError(found) Then Print #1, GetData(i)
     
Next i

Close #1
Wb1.Close SaveChanges:=False

End Sub

Function GetData(index As Integer)
Dim retVal As String

retVal = "Invoice Number: " + CStr(Worksheets(1).Cells(index, 1)) + ", Order Number: " + CStr(Worksheets(1).Cells(index, 3)) + ", Customer Address Key: " + CStr(Worksheets(1).Cells(index, 6)) + ", Customer Name: " + Worksheets(1).Cells(index, 7)
GetData = retVal
End Function """
